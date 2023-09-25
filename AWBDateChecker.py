import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from selenium.common.exceptions import TimeoutException
import os
import sys

# Initialize a WebDriver (Chrome, in this case)
driver = webdriver.Chrome()

# Get the directory where the Python script is located
script_directory = os.path.dirname(sys.argv[0])

# Set the path to your Excel file (assuming it's in the same directory)
excel_file_path = os.path.join(script_directory, 'LISTA MANIFIESTO+ETD.xlsx')

# Load the Excel file with write permissions (even if it's open)
workbook = openpyxl.load_workbook(excel_file_path, read_only=False)
sheet = workbook.active

# Get the year from cell I2 (assuming I2 contains the year)
year_cell = sheet['I2']
year = year_cell.value


# Start from the second row (row 2) assuming you have headers in row 1
for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
    awb_number = row[0]

    if awb_number is not None:  # Add this check
        # Navigate to the web page
        driver.get("http://isidora.aduana.cl/WebManifiestoAereo/Consultas/CON_GuiasAereasxMFTOA.jsp")

        # Find the input field and button elements
        input_field = driver.find_element(By.NAME, "EdNroGuia")
        consult_button = driver.find_element(By.NAME, "cmdBuscar")

        # Input the AWB number
        input_field.send_keys(awb_number)

        # Click the "Consultar" button
        consult_button.click()

        try:
            # Wait for the date to appear (adjust the wait time as needed)
            wait = WebDriverWait(driver, 10)
            date_element = wait.until(EC.presence_of_element_located((By.XPATH, '//td[contains(text(), "20")][contains(text(), "-")]')))

            # Extract the date text from the web and write it to column D (YYYY-MM-DD format)
            date_text = date_element.text.strip()
            date_text_parts = date_text.split()
            formatted_date = date_text_parts[0] if len(date_text_parts) > 0 else ""
            cell_d = sheet.cell(row=row_index, column=4)
            cell_d.value = formatted_date

            # Check if the date in column B is already a datetime object
            date_b = sheet.cell(row=row_index, column=2).value
            if isinstance(date_b, datetime):
                formatted_date_b = date_b.strftime("%Y-%m-%d")
            else:
                try:
                    date_b = datetime.strptime(date_b, "%d/%m/%Y")  # Format: "dd/mm/yyyy"
                    formatted_date_b = date_b.strftime("%Y-%m-%d")
                except ValueError:
                    formatted_date_b = ""
                    
            cell_c = sheet.cell(row=row_index, column=3)
            cell_c.value = formatted_date_b

        except TimeoutException:
            # Check if the message "No se encontraron Guías Aéreas" is present
            if "No se encontraron Guías Aéreas" in driver.page_source:
                sheet.cell(row=row_index, column=5, value="NO FOUND")
            else:
                sheet.cell(row=row_index, column=5, value="ERROR")

# Compare the dates in columns C and D
for row in sheet.iter_rows(min_row=2, max_col=4):
    date_c = row[2].value
    date_d = row[3].value

    if date_c != date_d:
        sheet.cell(row=row[0].row, column=5, value="YES")
    else:
        sheet.cell(row=row[0].row, column=5, value="NO")

# Save the Excel file
workbook.save(excel_file_path)

# Close the WebDriver
driver.quit()
