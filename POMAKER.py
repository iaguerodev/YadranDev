import win32com.client
import pyautogui
import pyperclip
import time
import pygetwindow as gw
import openpyxl
import os
from datetime import datetime, timedelta
import re
from openpyxl import load_workbook
import pandas as pd


## I-AGUERO ## BOT FRESH USA ##

def clean_and_format_value(value):
    if isinstance(value, str):
        cleaned_value = value.replace('$', '').replace('/', '').replace('cs', '').replace('.', ',').strip()
        return cleaned_value
    else:
        # If the value is not a string, return it as is
        return value



# Open the Excel file
excel_file_path = "PO.xlsx"  # Change this if your Excel file has a different name
workbook = openpyxl.load_workbook(excel_file_path)

# Select the appropriate sheet, assuming it's the first sheet
sheet = workbook.active

# Access and manipulate the 'PRECIO' column (assuming it's in column K)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=11, max_col=11):
    for cell in row:
        # Perform the desired manipulations
        cleaned_value = clean_and_format_value(cell.value)

        # Update the cell value with the cleaned value
        cell.value = cleaned_value

# Save the changes
workbook.save(excel_file_path)

    


# Function to get the FF code based on FF name
def get_ff_code(ff_name):
    ff_name = ff_name.upper()  # Convert to uppercase for consistent matching
    ff_codes = {
        'ALFA': '76272345-K',
        'AGILITY': '76408000-9',
        'ANDES': '76788050-2',
        'A&A': '78171100-4'
        # Add more mappings as needed
    }
    return ff_codes.get(ff_name, None)


# Function to check if the active window's title contains a specific substring
def is_window_title_containing(substring):
    active_window = gw.getActiveWindow()
    if substring in active_window.title:
        return True
    return False


# Function to refocus on sap window
def focus_sap_window(window_title):
    sap_window = gw.getWindowsWithTitle(window_title)
    if sap_window:
        sap_window[0].activate()

# Initialize SAP GUI Scripting
SapGuiAuto = win32com.client.GetObject("SAPGUI")
application = SapGuiAuto.GetScriptingEngine
connection = application.Children(0)
session = connection.Children(0)

# Check if a connection to SAP GUI is established
if session.Children.Count < 0:
    session = application.CreateObject("Session")

# Set the working pane dimensions
#Enter transaction
session.findById("wnd[0]").resizeWorkingPane(118, 27, 0)  # Use 0 instead of False
session.findById("wnd[0]/tbar[0]/okcd").text = "/nVA01"
session.findById("wnd[0]").sendVKey(0)


#Fill transaction with fresh 
session.findById("wnd[0]/usr/ctxtVBAK-AUART").text = "ZRCE"
session.findById("wnd[0]/usr/ctxtVBAK-VKORG").setFocus()
session.findById("wnd[0]/usr/ctxtVBAK-VKORG").caretPosition = 0

session.findById("wnd[0]/usr/ctxtVBAK-VKORG").text = "YD01"
session.findById("wnd[0]/usr/ctxtVBAK-VTWEG").text = "02"
session.findById("wnd[0]/usr/ctxtVBAK-SPART").text = "01"
session.findById("wnd[0]").sendVKey(0)

#Fill with client
# Set "10016" in field KUAGV-KUNNR
session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/subPART-SUB:SAPMV45A:4701/ctxtKUAGV-KUNNR").text = "10016"
# Set "10016" in field KUWEV-KUNNR
session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/subPART-SUB:SAPMV45A:4701/ctxtKUWEV-KUNNR").text = "10016"



#PO extraction
excel_file_path = os.path.join(os.path.dirname(__file__), "PO.xlsx")
try:
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active
    full_value = str(worksheet["A2"].value)
    workbook.close()
except Exception as e:
    print(f"Error reading data from Excel: {e}")
    full_value = ""

# Extract the first 5 numbers from the right to left
first_5_numbers = full_value[-5:]

# Set the first 5 numbers in the "VBKD-BSTKD" field ( PO number )
session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/txtVBKD-BSTKD").text = first_5_numbers


#DATE ( FECHA CARGA )
try:
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active
    day_of_week = worksheet["B2"].value
    workbook.close()
except Exception as e:
    print(f"Error reading data from Excel: {e}")
    day_of_week = ""

# Define a dictionary to map day names to date offsets (1 for "Lunes," 2 for "Martes," etc.)
day_to_offset = {
    "Lunes": 1,
    "Martes": 2,
    "Miercoles": 3,
    "Jueves": 4,
    "Viernes": 5,
    "Sabado": 6,
    "Domingo": 7,
}



# Get the current date
current_date = datetime.now()

# Find the offset of the specified day within the current week
current_weekday = current_date.isoweekday()
days_until_next_weekday = (day_to_offset[day_of_week] - current_weekday) % 7

# Calculate the date for the next week
next_week_date = current_date + timedelta(days=days_until_next_weekday)
if next_week_date <= current_date:
    next_week_date += timedelta(weeks=1)

# Format the date as "dd.mm.yyyy"
next_week_date_formatted = next_week_date.strftime("%d.%m.%Y")

# Set the calculated date in the "RV45A-KETDAT" field
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/ssubHEADER_FRAME:SAPMV45A:4440/ctxtRV45A-KETDAT").text = next_week_date_formatted

#INPUT INCOTERM -- USA DOBLE SLASH
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/ssubHEADER_FRAME:SAPMV45A:4440/ctxtVBKD-INCO1").text = "CPT"

# Load Excel data using openpyxl - INCOTERM
try:
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active
    cell_value = worksheet['H2'].value
    workbook.close()
except Exception as e:
    print(f"Error reading data from Excel: {e}")
    cell_value = ""

# Set the value in the SAP field VBKD-INCO2_L ( INCOTERM FIELD )
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/ssubHEADER_FRAME:SAPMV45A:4440/ctxtVBKD-INCO2_L").text = cell_value

#Cambiar a cabecera
session.findById("wnd[0]").sendVKey(0)
session.findById("wnd[0]").sendVKey(0)

# Resize working pane and click "btnBT_HEAD"
session.findById("wnd[0]").resizeWorkingPane(118, 27, False)
session.findById("wnd[0]/usr/subSUBSCREEN_HEADER:SAPMV45A:4021/btnBT_HEAD").press()

# Select the "tabpT\07" tab - INTERLOCUTOR TAB
command = "=T\\07"  # Use the exact command "=T\07"
session.findById("wnd[0]/tbar[0]/okcd").text = command
session.findById("wnd[0]/tbar[0]/btn[0]").press()  # Simulate pressing the Enter


# INTERLOCUTOR -- Set the FF code in the SAP field
# Load Excel data
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07/ssubSUBSCREEN_BODY:SAPMV45A:4352/subSUBSCREEN_PARTNER_OVERVIEW:SAPLV09C:1000/tblSAPLV09CGV_TC_PARTNER_OVERVIEW/ctxtGVS_TC_DATA-REC-PARTNER_EXT[1,5]").text = "77269270-6"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07/ssubSUBSCREEN_BODY:SAPMV45A:4352/subSUBSCREEN_PARTNER_OVERVIEW:SAPLV09C:1000/tblSAPLV09CGV_TC_PARTNER_OVERVIEW/ctxtGVS_TC_DATA-REC-PARTNER_EXT[1,5]").setFocus
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07/ssubSUBSCREEN_BODY:SAPMV45A:4352/subSUBSCREEN_PARTNER_OVERVIEW:SAPLV09C:1000/tblSAPLV09CGV_TC_PARTNER_OVERVIEW/ctxtGVS_TC_DATA-REC-PARTNER_EXT[1,5]").caretPosition = 0
session.findById("wnd[0]").sendVKey (0)

excel_file = 'PO.xlsx'
worksheet = workbook.active
ff_name = worksheet['O2'].value
workbook.close()

# Get the FF code
ff_code = get_ff_code(ff_name)
print(ff_code)

if ff_code:
    # Set the key value to "Z3" in the combo box
    session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07/ssubSUBSCREEN_BODY:SAPMV45A:4352/subSUBSCREEN_PARTNER_OVERVIEW:SAPLV09C:1000/tblSAPLV09CGV_TC_PARTNER_OVERVIEW/cmbGVS_TC_DATA-REC-PARVW[0,6]").key = "Z3"
    
    # Set the FF code in the text field
    session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\07/ssubSUBSCREEN_BODY:SAPMV45A:4352/subSUBSCREEN_PARTNER_OVERVIEW:SAPLV09C:1000/tblSAPLV09CGV_TC_PARTNER_OVERVIEW/ctxtGVS_TC_DATA-REC-PARTNER_EXT[1,6]").text = ff_code
else:
    print(f"FF code not found for {ff_name}")

# Select the "tabpT\11" tab - DATOS ADICIONALES A
command = "=T\\11"  
session.findById("wnd[0]/tbar[0]/okcd").text = command
session.findById("wnd[0]/tbar[0]/btn[0]").press()  # Simulate pressing the Enter

#Datos adicionales A - complete info
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBAK-KVGR1").key = "001"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBAK-KVGR2").key = "003"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBAK-KVGR3").key = "003"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBAK-KVGR4").key = "003"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBAK-KVGR5").key = "001"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBKD-KDKG1").key = "02"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBKD-KDKG2").key = "02"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBKD-KDKG3").key = "02"
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\11/ssubSUBSCREEN_BODY:SAPMV45A:4309/cmbVBKD-KDKG3").setFocus

# Select the "tabpT\12" tab - DATOS ADICIONALES B
command = "=T\\12"  
session.findById("wnd[0]/tbar[0]/okcd").text = command
session.findById("wnd[0]/tbar[0]/btn[0]").press()  # Simulate pressing the Enter


##  DATOS ADICIONALES B
#ETD
# Calculate the date for the next week
next_week_date = current_date + timedelta(days=days_until_next_weekday)
if next_week_date <= current_date:
    next_week_date += timedelta(weeks=1)

# Add one day to the calculated date
next_week_date += timedelta(days=1)

# Format the date as "dd.mm.yyyy"
next_week_date_formatted = next_week_date.strftime("%d.%m.%Y")

# Campo ETD
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/ctxtVBAK-ZFETD").text = next_week_date_formatted
# Campo FECHA CERT
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/ctxtVBAK-ZUBCE").text = next_week_date_formatted

# CAMPO ETA
# Calculate the date for the next week
next_week_date = current_date + timedelta(days=days_until_next_weekday)
if next_week_date <= current_date:
    next_week_date += timedelta(weeks=1)

# Add one day to the calculated date
next_week_date += timedelta(days=2)

# Format the date as "dd.mm.yyyy"
next_week_date_formatted = next_week_date.strftime("%d.%m.%Y")

# Fecha Arribo
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/ctxtVBAK-ZFETA").text = next_week_date_formatted

# Peso Objetivo
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/ctxtVBAK-ZPESO_OBJETIVO").text = "NO"

# Puerto Descarga
session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/cmbVBAK-ZEMBA").key = "AEROPUERTO C.A.M.B"

# Puerto Destino - Leer excel y pega en SAP FIELD "Puerto Destino"
try:
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active
    destination = worksheet["H2"].value
    workbook.close()
    print(f"Read destination from Excel: {destination}")
except Exception as e:
    print(f"Error reading data from Excel: {e}")
    destination = ""

# Set the SAP field's key based on the Excel value
if destination == "Miami":
    session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/cmbVBAK-ZDEST").key = "MIAMI"
elif destination == "LAX":
    session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/cmbVBAK-ZDEST").key = "LOS ANGELES"
elif destination == "JFK":
    session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_HEAD/tabpT\\12/ssubSUBSCREEN_BODY:SAPMV45A:4312/sub8309:SAPMV45A:8309/cmbVBAK-ZDEST").key = "NEW YORK"
else:
    print("Unknown destination from Excel")

time.sleep(1)

#Volver a principal

# Simulate pressing the F3 key
session.findById("wnd[0]").sendVKey(3)
time.sleep(3)

########################     INGRESO MATERIAL     ####################################
######################################################################################

# Function for finding the first match index

## MATERIAL SELECTION ##
# Define the criteria and corresponding values

criteria_and_values = {

    # Fillets TD PBI x 16 Descamado
    ("D PBI / scl-OFF", "1-2", 16): "12000678",
    ("D PBI / scl-OFF", "2-3", 16): "12000679",
    ("D PBI / scl-OFF", "3-4", 16): "12000680",
    ("D PBI / scl-OFF", "4-5", 16): "12000681",

    ("D PBI / scl-OFF", "2.0-2.5", 16): "12000970",
    ("D PBI / scl-OFF", "2.5-3.0", 16): "12000971",
    ("D PBI / scl-OFF", "3.0-3.5", 16): "12000973",
    ("D PBI / scl-OFF", "3.5-4.0", 16): "12000974",

    # fillets TD PBO x 16 Descamado
    ("D / scl-OFF", "1-2", 16): "12000314",
    ("D / scl-OFF", "2-3", 16): "12000187",
    ("D / scl-OFF", "3-4", 16): "12000192",
    ("D / scl-OFF", "4-5", 16): "12000186",

    # fillets TD PBO x 36 Descamado
    ("D / scl-OFF", "1-2", 36): "12000312",
    ("D / scl-OFF", "2-3", 36): "12000737",
    ("D / scl-OFF", "3-4", 36): "12000193",
    ("D / scl-OFF", "4-5", 36): "12000207",

    # fillets TD PBO x 36 Descamado
    ("D", "1-2", 36): "12000279",
    ("D", "2-3", 36): "12000188",
    ("D", "3-4", 36): "12000184",
    ("D", "4-5", 36): "12000189",

    # fillets 1-4
    ("D / scl-OFF", "1-4", 36): "12000914",
    ("D", "1-4", 36): "12000913",

    # fillets x 12 PBO con escama
    ("D", "1-2", 12): "12000877",
    ("D", "2-3", 12): "12000200",
    ("D", "3-4", 12): "12000199",
    ("D", "4-5", 12): "12000309",

    # fillets x 12 PBO SIN escama
    ("D / scl-OFF", "1-2", 12): "12000915",
    ("D / scl-OFF", "3-4", 12): "12000373",
    ("D / scl-OFF", "4-5", 12): "12001125",

    # TRIM E x 12
    ("E", "2-3", 12): "12000307",
    ("E", "3-4", 12): "12000330",

    # TRIM E x 36
    ("E", "2-3", 36): "12000250",
    ("E", "3-4", 36): "12000190",

    # HON 
    ("HG", "10-12", 35): "12000458",
    ("HON", "16+", 55): "12000387",
    ("HON", "14-16", 55): "12000368",
    ("HON", "8-10", 55): "12000687",
    ("HON", "8-12", 25): "12000912",

    # Portions
    ("Porc. Bias & SwS", "7 oz", 11): "12000882",
    ("Porc. Bias & SwS", "9 oz", 11): "12000881",
    ("Porciones C/Piel scl-OFF", "5-6 oz RL", 21): "12000396",
    ("Mignon FIX", "min 6.00 oz", 5) : "12000210",

    # SCL-ON
    ("D / scl-ON", "1-2", 16): "12000901",
    ("D / scl-ON", "2-3", 16): "12000755",
    ("D / scl-ON", "3-4", 16): "12000756",
    ("D / scl-ON", "4-5", 16): "12000754",

    ("D PBI / scl-ON", "1-2", 16): "12000900",
    ("D PBI / scl-ON", "2-3", 16): "12000289",
    ("D PBI / scl-ON", "3-4", 16): "12000290",
    ("D PBI / scl-ON", "4-5", 16): "12000291",

    ("D / scl-ON", "2-3", 12): "12000904",

    ("E  -  PBI - Clean Cut", "2-3", 36): "12001123",
    ("E  -  PBI - Clean Cut", "3-4", 36): "12001124"

}

match_found = False  # Initialize a flag to track if a match is found

def find_first_match_index(excel_file_path, criteria_and_values):
    try:
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        worksheet = workbook.active

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True, min_col=4, max_col=6), start=1):
            criteria = row
            if criteria in criteria_and_values:
                return row_index

        # Return None if no match is found
        return None
    except Exception as e:
        print(f"Error finding first match index: {e}")
        return None
    finally:
        workbook.close()


# Function to process a material slot - SELECCIÓN MATERIAL
def process_material_slot(session, criteria_and_values, material_slot_index, material_row):
    try:
        material = criteria_and_values.get(tuple(material_row))
        if material is not None:
            # Set the SAP field text with the material value
            material_field_path = f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG/ctxtRV45A-MABNR[1,{material_slot_index - 1}]"
            session.findById(material_field_path).setFocus()
            session.findById(material_field_path).text = material
            print(f"Material code input in SAP for material slot {material_slot_index}: {material}")

    except Exception as e:
        print(f"Error processing material slot {material_slot_index}: {e}")

try:
    excel_file_path = os.path.join(os.path.dirname(__file__), "PO.xlsx")
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active

    # Get all rows with matching criteria
    matching_rows = [row for row in worksheet.iter_rows(min_row=2, values_only=True, min_col=4, max_col=6) if tuple(row) in criteria_and_values]

    if not matching_rows:
        print("No matching criteria found in Excel.")
    else:
        print(f"{len(matching_rows)} matching criteria found in Excel.")

        # Process each material slot
        for material_slot_index, material_row in enumerate(matching_rows, start=1):
            process_material_slot(session, criteria_and_values, material_slot_index, material_row)

except Exception as e:
    print(f"Error reading data from Excel: {e}")
finally:
    workbook.close()

prices_column = [row[10] for row in worksheet.iter_rows(min_row=2, values_only=True)]

# Load Excel Data
try:
    po_workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    po_worksheet = po_workbook.active
    cajas_column = [row[6] for row in po_worksheet.iter_rows(min_row=2, values_only=True)]
    

    # Process each material slot
    for material_slot_index, (material_row, cajas_text, price_text) in enumerate(zip(matching_rows, cajas_column, prices_column), start=1):
        process_material_slot(session, criteria_and_values, material_slot_index, material_row)


        # Set Text in SAP for "CAJAS" field for RV45A-KWMENG[2,{material_slot_index}]
        session_path = f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG/txtRV45A-KWMENG[2,{material_slot_index - 1}]"
        session.findById(session_path).text = cajas_text

        # Set Material Features
        session.findById(f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG/ctxtVBAP-WERKS[11,{material_slot_index - 1}]").text = "2001"
        session.findById(f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG/cmbVBAP-VKAUS[6,{material_slot_index - 1}]").key = "001"

        # Get the material code from the criteria_and_values dictionary
        material = criteria_and_values.get(tuple(material_row), "")
        
        # Set VBKD-KONDA based on material code
        session_path = f"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG/cmbVBKD-KONDA[13,{material_slot_index - 1}]"
        try:
            # Debug: Print the material code and the decision
            print(f"Material Code: {material}")

            # Set the value directly based on the condition
            if material in ["12000210", "12000265", "12000201", "12000202", "12000882", "12000881"]:
                session.findById(session_path).key = "01"
                print("Setting VBKD-KONDA to '01'")
            else:
                session.findById(session_path).key = "03"
                print("Setting VBKD-KONDA to '03'")
        except Exception as e:
            print(f"Error setting VBKD-KONDA text: {e}")


        # SUB POSICION ENTER
        
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG").getAbsoluteRow(material_slot_index - 1).selected = 'True'
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\01/ssubSUBSCREEN_BODY:SAPMV45A:4400/subSUBSCREEN_TC:SAPMV45A:4900/tblSAPMV45ATCTRL_U_ERF_AUFTRAG/ctxtRV45A-MABNR[1,0]").showContextMenu()
        session.findById("wnd[0]").sendVKey (2)

        # Select the "tabpT\05" tab - Condiciones TAB
        command = "=T\\05"  # Use the exact command "=T\05"
        session.findById("wnd[0]/tbar[0]/okcd").text = command
        session.findById("wnd[0]/tbar[0]/btn[0]").press()


        # price

        # Set Precio Finder
        # Assuming K column in Excel represents the prices
        prices_column = [row[10] for row in po_worksheet.iter_rows(min_row=2, values_only=True)]
        precio_finder_path = f"wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\05/ssubSUBSCREEN_BODY:SAPLV69A:6201/tblSAPLV69ATCTRL_KONDITIONEN/txtKOMV-KBETR[3,0]"
        session.findById(precio_finder_path).setFocus()
        session.findById(precio_finder_path).text = price_text
        session.findById(precio_finder_path).caretPosition = 16
        session.findById("wnd[0]").sendVKey(0)


        # Select the "tabpT\13" tab - Datos Adicionales A TAB
        command = "=T\\13"  # Use the exact command "=T\13"
        session.findById("wnd[0]/tbar[0]/okcd").text = command
        session.findById("wnd[0]/tbar[0]/btn[0]").press()

        # Datos Adicionales A Tab
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\13/ssubSUBSCREEN_BODY:SAPMV45A:4459/cmbVBAP-MVGR1").key = "023"
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\13/ssubSUBSCREEN_BODY:SAPMV45A:4459/cmbVBAP-MVGR2").key = "000"
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\13/ssubSUBSCREEN_BODY:SAPMV45A:4459/cmbVBAP-MVGR3").key = "000"
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\13/ssubSUBSCREEN_BODY:SAPMV45A:4459/cmbVBAP-MVGR4").key = "000"
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\13/ssubSUBSCREEN_BODY:SAPMV45A:4459/cmbVBAP-MVGR5").key = "001"
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\13/ssubSUBSCREEN_BODY:SAPMV45A:4459/cmbVBAP-MVGR5").setFocus
        session.findById("wnd[0]/usr/tabsTAXI_TABSTRIP_ITEM/tabpT\\14").select
        session.findById("wnd[0]/tbar[0]/btn[3]").press()

finally:
    po_workbook.close()